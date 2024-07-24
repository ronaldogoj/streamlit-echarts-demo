
import time

import streamlit as st

import pandas as pd
from pandas import DataFrame
import base64
from io import BytesIO
import xlsxwriter
from openpyxl import load_workbook


class Rateio:
    def __init__(self):
        self.num_apartamentos: int = 84
        self.cota_minima_individual: int = 15
        self.tra: float = 0.0000
        self.esgoto: int = 1
        self.entrada: DataFrame = pd.DataFrame()
        self.config_conta: DataFrame = pd.DataFrame()
        self.cota_geral: int = self.num_apartamentos * self.cota_minima_individual
        self.faixas: list = [1,2,3,4,5]
        self.multiplicador: list = [1.0, 2.5, 3.1, 6.0, 8.0]
        self.v_m3_agua: list = []
        self.v_tarifa_agua: list = []
        self.total_geral: int = 0
        self.taxa: float = 0.0
        self.total_individual: int = 0
        self.total_comum: int = 0
        self.aloc_min_comum_f1: int = 1
        self.aloc_max_comum_f1: int = 200
        self.cota_min_ind: float = 0.0
        self.cons_ind_faixas_medicao: list = [0.0, 0.0, 0.0, 0.0, 0.0]
        self.cons_ind_faixas_geral: list = []
        self.valor_cota_min_ind: float = 0.0
        self.valor_total_comum: float = 0.0
        self.unidade = 0
        self.arquivo = None

    def ler_arquivo_excel_entrada(self):
        pass

    def menu_lateral(self):
        # Menu lateral
        with st.sidebar:
            st.sidebar.header("Configurações")
            self.num_apartamentos = st.sidebar.number_input("Número de Apartamentos", min_value=1, value=self.num_apartamentos, step=1)
            self.cota_minima_individual = st.sidebar.number_input("Cota Mínima Individual", min_value=1, value=self.cota_minima_individual, step=1)
            #self.tra = st.sidebar.number_input("TRA", min_value=0.0000, value=self.tra, step=0.0001, format="%.4f")
            #self.total_geral = st.sidebar.number_input("Total Geral", min_value=0, value=self.total_geral, step=1)
            #self.taxa = st.sidebar.number_input("Taxa", min_value=0.00, value=self.taxa, step=0.01, format="%.2f")
            #self.aloc_max_comum_f1 = st.sidebar.number_input("Alocação Max F1 Comum", min_value=1, value=self.aloc_max_comum_f1, step=1)

            # slider_val = st.slider("Form slider")


    def pagina_principal(self):
        uploaded_file = st.file_uploader("Choose a file", type=["csv", "txt", "xlsx"])

        if uploaded_file is not None:
            st.subheader("Resultado:")
            file_extension = uploaded_file.name.split(".")[-1]

            # Salvar o arquivo para uso posterior
            #self.arquivo = uploaded_file

            # Check file type and read accordingly
            if file_extension.lower() in ["xls", "xlsx"]:
                # Ler o arquivo de entrada
                self.entrada = pd.read_excel(uploaded_file, sheet_name="Consumo", engine='openpyxl')
                self.config_conta = pd.read_excel(uploaded_file, sheet_name="Conta", engine='openpyxl')
            else:
                st.error(f"Tipo de arquivo incompatível: {file_extension}")
                return

            # Display the DataFrame
            #st.write(self.entrada[['Unidade', 'consumo']])

            # Show result in a popup
            # st.success("Arquivo enviado com sucesso!")

            # Chamar a verificação do arquivo
            self.verif_arquivo()

    def exibir_resumo(self, titulo, valor, exib_res=False, tipo="text"):
        with st.sidebar:
            if exib_res:
                st.sidebar.header("Resumo")
            st.text_input(titulo, valor, disabled=True)



    def calcular_rateio(self):
        def max_aloc_faixa(fai, tc, f1_c_orig, v_t_a, cg):
            # A alocação da faixa 1 é fixa, mas das faixas posteriores pode mudar um pouco dada a demanda, podendo ocorrer empréstimo entre os tipos
            tipos = {0: 'comum', 1: 'individual'}
            res = []
            alocado = {0: 0.0, 1: 0.0}

            for x in fai:
                sobrou = cg
                faltou = 0.0
                # Voltar para os limites originais sempre que iniciar uma nova fase
                f1_c = f1_c_orig.copy()
                # Pegar emprestado - definir o limite da faixa 2 em diante (f1_c)
                if x > 0:  # testar desde a faixa 1, pois pode acontecer, mas falta ver se funciona
                    # Ver se algum tipo vai precisar de mais um pouco, além do seu limite inicial
                    # Tipo Comum
                    precisa = 0.0
                    sobra = 0.0
                    sobra_tipo_anterior = 0.0
                    for key, value in tipos.items():
                        emprestar = 0.0
                        pendente = tc[key] - alocado[key]
                        limite = f1_c[key]
                        sobra = 0.0 if pendente > limite else limite - pendente
                        # Verificar se o tipo anterior precisa de algo e se tem algo sobrando do atual
                        if precisa > 0 and sobra > 0:
                            # O tipo anterior precisa e temos algum valor sobrando
                            if sobra >= precisa:
                                # Se tem sobrando o suficiente
                                emprestar = precisa
                            else:
                                # Não tem o total, mas tem alguma coisa
                                emprestar = sobra
                            # Já que vamos emprestar, atualizar os limites da faixa atual
                            f1_c[key - 1] += emprestar
                            f1_c[key] -= emprestar
                        # Verifica se o tipo atual precisa de empréstimo
                        if pendente > limite:
                            precisa = pendente - limite
                            # Como ver a sobra do tipo anterior
                            # Verificar se tem sobra do tipo anterior
                            if precisa > 0 and sobra_tipo_anterior > 0:
                                # O tipo atual precisa e existe sobra do tipo anterior
                                if sobra_tipo_anterior >= precisa:
                                    # Se tem sobrando o suficiente
                                    emprestar = precisa
                                else:
                                    # Não tem o total, mas tem alguma coisa
                                    emprestar = sobra_tipo_anterior
                                # Já que vamos emprestar, atualizar os limites da faixa atual
                                f1_c[key - 1] -= emprestar
                                f1_c[key] += emprestar
                        # Salvar a sobra do tipo atual
                        sobra_tipo_anterior = sobra
                # Definir os valores por faixa
                for key, value in tipos.items():
                    pendente = tc[key] - alocado[key]
                    limite = f1_c[key]
                    if pendente > limite:
                        qtd_alocar = f1_c[key]
                        faltou = tc[key] - alocado[key] - qtd_alocar
                    else:
                        qtd_alocar = tc[key] - alocado[key]
                        faltou = 0.0
                    alocado[key] += qtd_alocar
                    sobrou -= qtd_alocar
                    valor_faixa = round(qtd_alocar * v_t_a[x - 1], 2)
                    linha = {}
                    linha['tipo'] = value
                    linha['faixa'] = x
                    linha['pendente'] = pendente
                    linha['limite'] = limite
                    linha['qtd'] = qtd_alocar
                    linha['valor'] = valor_faixa
                    linha['sobrou'] = sobrou
                    linha['faltou'] = faltou
                    res.append(linha)
                # Criar o DF
                df_res = pd.DataFrame(res)
                # Gerar as saídas antigas
                val_f_c = []
                val_f_i = []
                cons_c_f_g = []
                cons_i_f_g = []
                valor_t_c = 0.0
                valor_t_i = 0.0
                for index, row in df_res.iterrows():
                    if row['tipo'] == 'comum':
                        val_f_c.append([row['faixa'], row['qtd'], row['valor']])
                        cons_c_f_g.append(row['qtd'])
                        valor_t_c += row['valor']
                    else:
                        val_f_i.append([row['faixa'], row['qtd'], row['valor']])
                        cons_i_f_g.append(row['qtd'])
                        valor_t_i += row['valor']
            return df_res, val_f_c, val_f_i, cons_c_f_g, cons_i_f_g, valor_t_c, valor_t_i

        # Função para calcular a tarifa
        def calcular_tarifa():
            # TARIFA Calculada
            df_individual = pd.DataFrame()
            df_individual['faixa'] = self.faixas
            df_individual['medicao'] = self.cons_ind_faixas_medicao
            df_individual['geral'] = self.cons_ind_faixas_geral

            for f in self.faixas:
                df_individual[f'tarifa{f}'] = 0.0
            df_individual['sobrou'] = 0.0

            sobrou = 0.0
            faixa = 1
            for index, row in df_individual.iterrows():
                con = row['medicao']
                lim = row['geral']

                for t in range(faixa, 6):
                    disp = sobrou if sobrou > 0 else lim

                    if row['faixa'] == 5:
                        # Estamos na última faixa, então não podemos ter sobra
                        alocar_tar_orig = disp
                        # Resolver o problema de aumento de valor após ter sobra na última faixa
                        # Atualizar o valor de medição com base na quantidade de m3 alocados
                        df_individual['medicao'].at[index] = alocar_tar_orig
                    else:
                        alocar_tar_orig = disp if con > disp else con


                    sobrou = disp - alocar_tar_orig
                    con -= alocar_tar_orig

                    df_individual[f'tarifa{faixa}'].at[index] = alocar_tar_orig

                    if sobrou > 0 or con <= 0:
                        df_individual['sobrou'].at[index] = sobrou
                        break
                    else:
                        faixa += 1



            # Descobrir a diferença de valor da faixa 1 para o consumo individual (repasse cota mínima individual
            valor_faixas_medicao = [sum([round(df_individual[f'tarifa{x}'].at[y - 1] * self.v_tarifa_agua[x - 1], 4) for x in self.faixas]) for y in self.faixas]
            valor_faixa1_medicao = valor_faixas_medicao[0]
            valor_faixa1_conta = val_faixas_ind[0][2]
            dif_faixa1 = round(valor_faixa1_conta - valor_faixa1_medicao, 4)
            df_individual['val_faix_med'] = valor_faixas_medicao
            # Percentual de valor das demais faixas
            df_individual['proporcao_valor'] = 0.0
            df_individual['proporcao_valor'].loc[(df_individual['medicao'] > 0) & (df_individual.index > 0)] = (
                        df_individual['val_faix_med'] / df_individual['val_faix_med'].loc[
                    (df_individual['medicao'] > 0) & (df_individual.index > 0)].sum())
            # Definição da tarifa das faixas 2 em diante **** será a tarifa final *******
            df_individual['tarifa_valor'] = 0.0
            df_individual['tarifa_valor'].loc[df_individual['proporcao_valor'] > 0] = df_individual['val_faix_med'] / df_individual['medicao']
            # Novos valores para as faixas (já com a sobra da faixa 1 distribuído entre as demais faixas)
            df_individual['val_faix_med_atualizado'] = 0.0
            df_individual['val_faix_med_atualizado'].at[0] = round(valor_faixa1_conta, 2)
            df_individual['val_faix_med_atualizado'].loc[df_individual['proporcao_valor'] > 0] = round(
                df_individual['val_faix_med'] - (dif_faixa1 * df_individual['proporcao_valor']), 2)
            df_individual['val_faix_med_atualizado'].sum()
            # Definição da tarifa das faixas 2 em diante **** será a tarifa final *******
            df_individual['tarifa_atualiz'] = 0.0
            df_individual['tarifa_atualiz'].loc[df_individual['proporcao_valor'] > 0] = df_individual['val_faix_med_atualizado'] / df_individual['medicao']
            # valor com base na tarifa original
            df_individual['tarifa_original'] = self.v_tarifa_agua
            df_individual['val_faix_med_tar_original'] = 0.0
            df_individual['val_faix_med_tar_original'].loc[df_individual.index > 0] = df_individual['medicao'] * df_individual['tarifa_original']
            # Percentual de valor das demais faixas com base na tarifa original
            df_individual['proporcao_valor_orig'] = 0.0
            df_individual['proporcao_valor_orig'].loc[(df_individual['medicao'] > 0) & (df_individual.index > 0)] = (
                        df_individual['val_faix_med_tar_original'] / df_individual['val_faix_med_tar_original'].loc[
                    (df_individual['medicao'] > 0) & (df_individual.index > 0)].sum())
            # valor com base na proporção original
            df_individual['val_faix_med_tar_original_ajus'] = 0.0
            df_individual['val_faix_med_tar_original_ajus'].loc[df_individual.index > 0] = (valor_total_indiv - valor_faixa1_conta) * df_individual['proporcao_valor_orig']
            # Definição da tarifa das faixas 2
            df_individual['tarifa_orig'] = 0.0
            df_individual['tarifa_orig'].loc[df_individual['proporcao_valor'] > 0] = df_individual['val_faix_med_tar_original_ajus'] / df_individual['medicao']
            # Proporção média
            df_individual['proporcao_media'] = 0.0
            df_individual['proporcao_media'] = (df_individual['proporcao_valor'] + df_individual['proporcao_valor_orig']) / 2
            # valor com base na proporção média
            df_individual['val_faix_med_tar_media'] = 0.0
            df_individual['val_faix_med_tar_media'].loc[df_individual.index > 0] = ( valor_total_indiv - valor_faixa1_conta) * df_individual['proporcao_media']
            # Definição da tarifa das faixas 2 em diante **** será a tarifa final *******
            df_individual['tarifa_media'] = 0.0
            df_individual['tarifa_media'].loc[df_individual['proporcao_media'] > 0] = df_individual['val_faix_med_tar_media'] / df_individual['medicao']
            # Tarifa que vai ser usada é a média
            df_individual['tarifa'] = df_individual['tarifa_media']
            return df_individual, valor_faixa1_conta

        # Preparar o rateio
        def preparar_rateio(df_individual):
            # Preparar o valor de rateio para as unidades
            df_rateio = self.entrada[['consumo']].copy()
            df_rateio.index = self.entrada['Unidade']
            for x in range(1, 6):
                df_rateio[f'qtd_faixa{x}'] = 0.0
                df_rateio[f'val_faixa{x}'] = 0.0
            df_rateio['val_faixa1'] = self.valor_cota_min_ind
            df_rateio['qtd_faixa1'] = df_rateio['consumo'].apply(lambda x: self.cota_min_ind if x > self.cota_min_ind else x)
            for index, row in df_rateio.iterrows():
                con = row['consumo']
                con -= self.cota_min_ind
                if con > 0:
                    for x in range(2, 6):
                        # Verificar se estamos na última faixa
                        if x == 5:
                            # Última faixa
                            q_m3 = con
                        else:
                            # Não é a última faixa
                            q_m3 = self.cota_min_ind if con > self.cota_min_ind else con
                        con -= q_m3
                        df_rateio[f'qtd_faixa{x}'].at[index] = q_m3
                        df_rateio[f'val_faixa{x}'].at[index] = round(q_m3 * df_individual['tarifa'].at[x - 1], 2)
                        if con == 0:
                            break
            df_rateio['val_individual'] = df_rateio['val_faixa1'] + df_rateio['val_faixa2'] + df_rateio['val_faixa3'] + df_rateio['val_faixa4'] + df_rateio['val_faixa5']
            self.entrada['fracao_ideal'] = self.entrada['fracao_ideal'].apply(lambda x: round(x, 5))
            df_rateio['val_comum'] = 0.0
            df_rateio['val_comum'] = self.entrada['fracao_ideal'].tolist()
            df_rateio['val_comum'] = df_rateio['val_comum'].apply(lambda x: round(x * self.valor_total_comum, 2))
            # Valor final do rateio
            df_rateio['valor_final'] = df_rateio['val_individual'] + df_rateio['val_comum']
            df_rateio['valor_final'].sum()
            df_rateio['val_comum'].sum()
            df_resumo_final = df_rateio[['consumo', 'valor_final', 'val_individual', 'val_comum']]
            return df_rateio, df_resumo_final

        def real_br_money_mask(my_value):
            a = '{:,.2f}'.format(float(my_value))
            b = a.replace(',', 'v')
            c = b.replace('.', ',')
            return c.replace('v', '.')

        # Function to save all dataframes to one single excel
        def dfs_tabs(df_list, sheet_list, file_name):

            output = BytesIO()

            writer = pd.ExcelWriter(output, engine='xlsxwriter')
            for dataframe, sheet in zip(df_list, sheet_list):
                dataframe.to_excel(writer, sheet_name=sheet, startrow=0, startcol=0)
            writer.close()

            processed_data = output.getvalue()
            return processed_data


        def gerar_excel_formatado(df, m3_total, tra, sheet="Rateio"):
            # Carregue a planilha existente
            workbook = load_workbook('modelo.xlsx')

            # Escolha a planilha desejada (pode ser a ativa ou outra específica)
            sheet = workbook.active

            # Escreva os valores do DataFrame na planilha
            for row_index, row in df.reset_index(drop=True).iterrows():
                for col_index, value in enumerate(row):
                    print(f'Valor: {value} - {row_index} - {col_index}')
                    sheet.cell(row=row_index + 3, column=col_index + 2, value=value)

            # Valores com posição fixa
            sheet.cell(row=8, column=8, value=m3_total)
            sheet.cell(row=14, column=12, value=tra)

            # Save the workbook to binary data
            binary_data = BytesIO()
            workbook.save(binary_data)
            binary_data.seek(0)

            return binary_data.getvalue()

        def exibir_detalhes_unidade(unidade):



            consumo_unidade = df_resumo_final['consumo'].loc[df_resumo_final.index == unidade].values[0]

            st.write(f"Unidade -> {unidade}")
            st.write(f"Consumo da unidade -> {consumo_unidade}")
            st.write(f"Valor total da conta -> {df_resumo_final['valor_final'].loc[df_resumo_final.index == unidade].values[0]}")

            #valor_final_unidade = df_resumo_final['valor_final'].loc[df_resumo_final['Unidade'] == unidade]
            #valor_individual = df_resumo_final['val_individual'].loc[df_resumo_final['Unidade'] == unidade]
            #valor_comum = df_resumo_final['val_comum'].loc[df_resumo_final['Unidade'] == unidade]
            m3_f=0
            resultado_individual = []
            for f in range(1, 6):
                if consumo_unidade == 0:
                    break
                linha = {}

                if f == 5:
                    # Última faixa
                    consumo_faixa = consumo_unidade
                else:
                    # Não é a última faixa
                    consumo_faixa = consumo_unidade if consumo_unidade < self.cota_min_ind else self.cota_min_ind

                linha['Tipo'] = 'Individual'
                linha['Faixa'] = f

                linha['De'] = m3_f
                m3_f += self.cota_min_ind
                if f == 5:
                    linha['Até'] = float('inf')
                else:
                    linha['Até'] = m3_f

                linha['Consumo'] = consumo_faixa

                tarifa = df_individual['tarifa'].loc[df_individual['faixa'] == f].values[0]
                tarifa_original = df_individual['tarifa_original'].loc[df_individual['faixa'] == f].values[0]
                linha['Tarifa Original'] = real_br_money_mask(tarifa_original)
                if f > 1:
                    linha['Tarifa Calculada'] = real_br_money_mask(tarifa)
                    linha['Valor'] = real_br_money_mask(round(consumo_faixa * tarifa, 2))
                else:
                    linha['Tarifa Calculada'] = real_br_money_mask(0)
                    linha['Valor'] = real_br_money_mask(self.valor_cota_min_ind)

                consumo_unidade -= consumo_faixa

                resultado_individual.append(linha)

            # Comum
            linha = {}
            linha['Tipo'] = 'Comum'
            linha['Valor'] = df_resumo_final['val_comum'].loc[df_resumo_final.index == unidade].values[0]
            resultado_individual.append(linha)

            # Dataframe
            df_resultado_final = pd.DataFrame(resultado_individual)


            st.write(df_resultado_final)





        # Ajuste no DF



        # Calcular o consumo com base nas medições dos relógios, caso o consumo esteja zerado na planilha
        if self.entrada['consumo'].sum() == 0:
            self.entrada['consumo'] = self.entrada['depois'] - self.entrada['antes']

        # Preencher o consumo total e taxa com base no valor da planilha
        #df_info_conta = pd.read_excel(self.arquivo, sheet_name="Conta", engine='openpyxl')
        # Ler o valor da taxa a partir da coluna "Configurações" com o valor "Taxa"
        self.taxa = self.config_conta.loc[self.config_conta['Configurações'] == 'Taxa']['Valor'].values[0] + self.config_conta.loc[self.config_conta['Configurações'] == 'Juros']['Valor'].values[0] + self.config_conta.loc[self.config_conta['Configurações'] == 'Multa']['Valor'].values[0] + self.config_conta.loc[self.config_conta['Configurações'] == 'Outros']['Valor'].values[0]
        # Ler o valor do consumo total a partir da coluna "Configurações" com o valor "Faturado (m3)"
        self.total_geral = int(self.config_conta.loc[self.config_conta['Configurações'] == 'Faturado (m3)']['Valor'].values[0])
        #self.total_geral = df_info_conta.loc[df_info_conta['Configurações'] == 'Faturado (m3)']['Valor'].values[0]

        self.tra = self.config_conta.loc[self.config_conta['Configurações'] == 'TRA']['Valor'].values[0]
        self.tra = st.sidebar.number_input("TRA", min_value=0.0000, value=self.tra, step=0.0001, format="%.4f")

        self.v_m3_agua: list = [x * self.tra for x in self.multiplicador]
        self.v_tarifa_agua = [x + (x * self.esgoto) for x in self.v_m3_agua]

        self.total_geral = st.sidebar.number_input("Total Geral", min_value=0, value=self.total_geral, step=1)
        self.taxa = st.sidebar.number_input("Taxa", value=self.taxa, step=0.01, format="%.2f")



        # Verificar se temos a quantidade total de m3 consumidos no mês
        if self.total_geral == 0:
            # O valor não foi informado pelo usuário
            st.error("Faltou informar o consumo total em m3!")
            return
        
        # Total individual e total comum
        self.total_ind = self.entrada['consumo'].sum()
        self.total_comum = self.total_geral - self.total_ind
        
        # Atribuir ao consumo individual a diferença até chegar na cota mínima (1260)
        if self.total_ind + self.total_comum < self.cota_geral:
            self.total_ind = self.cota_geral - self.total_comum
        
        # Ajustar o valor do total geral (caso não alcance o mínimo que precisa)
        self.total_geral = self.total_geral if self.total_geral >= self.cota_geral else self.cota_geral

        # Verificar se temos o valor da taxa
        if self.taxa == 0.0:
            # O valor não foi preenchido
            st.error("Faltou informar a taxa!")
            return

        # Exibir resumo
        self.exibir_resumo("Total Individual", self.total_ind, True)
        self.exibir_resumo("Total Comum", self.total_comum)


        # Atualizar o slider
        self.aloc_max_comum_f1 = self.total_comum
        self.aloc_min_comum_f1 = 0 if self.total_ind >= (self.num_apartamentos * self.cota_minima_individual) else (self.num_apartamentos * self.cota_minima_individual) - self.total_ind

        valor_default_sugerido = int(self.config_conta.loc[self.config_conta['Configurações'] == 'Aloc Max Comum']['Valor'].values[0])
        valor_default_max_f1_consumo = valor_default_sugerido if self.aloc_max_comum_f1 > valor_default_sugerido and self.aloc_min_comum_f1 < valor_default_sugerido else self.aloc_min_comum_f1

        if self.aloc_max_comum_f1 != self.aloc_min_comum_f1:
            self.aloc_max_comum_f1 = st.sidebar.slider("Alocação Max F1 Comum", min_value=self.aloc_min_comum_f1,
                                                   max_value=int(self.aloc_max_comum_f1), value=valor_default_max_f1_consumo,
                                                   step=1)
        else:
            self.aloc_max_comum_f1 = st.sidebar.number_input("Alocação Max F1 Comum", min_value=self.aloc_min_comum_f1, max_value=int(self.aloc_max_comum_f1), value=valor_default_max_f1_consumo, step=1)

        #atualizar o valor do slider


        # Usar um valor definido para alocação máxima do consumo comum
        # Esse valor pode mudar
        f1_cons_com: int = self.aloc_max_comum_f1
        f1_cons_ind: int = self.cota_geral - f1_cons_com

        # Ver se podemos pegar emprestado
        df_resultado, val_faixas_com, val_faixas_ind, cons_com_faixas_geral, self.cons_ind_faixas_geral, self.valor_total_comum, valor_total_indiv = max_aloc_faixa(
            self.faixas, [self.total_comum, self.total_ind], [f1_cons_com, f1_cons_ind], self.v_tarifa_agua, self.cota_geral)
        self.valor_total_comum += self.taxa

        # Ver se a alocação na faixa 1 mudou
        # Se mudou, tem que atualizar a alocação da faixa 1, que foi definida anteriormente
        if cons_com_faixas_geral[0] != f1_cons_com:
            # Mudou... Tem que atualizar
            f1_cons_com: int = cons_com_faixas_geral[0]
            f1_cons_ind: int = self.cota_geral - f1_cons_com

        # Definir a proporção para cálculo da cota mínima individual
        # Proporção do consumo individual alocado na faixa 1
        prop_f1_ind = round(f1_cons_ind / self.cota_geral, 2)

        # Definição da cota mínima individual com base no percentual do consumo individual alocado na faixa 1
        self.cota_min_ind = self.cota_minima_individual * prop_f1_ind
        self.exibir_resumo("Cota mínima individual", self.cota_min_ind)

        # Valores globais por tipo de consumo
        # Área Comum
        faixas_com = [[(x - 1) * f1_cons_com, x * f1_cons_com] for x in self.faixas]
        faixas_com[-1][-1] = float("nan")
        # Individual
        faixas_ind = [[(x - 1) * f1_cons_ind, x * f1_cons_ind] for x in self.faixas]
        faixas_ind[-1][-1] = float("nan")

        # Consumo individual por faixa de Medição

        for e in self.entrada.index.tolist():
            alocado = 0
            con_t = self.entrada['consumo'].at[e]
            for f in self.faixas:
                val = self.cota_min_ind if con_t - alocado > self.cota_min_ind else con_t - alocado
                alocado += val
                self.cons_ind_faixas_medicao[f - 1] = round(self.cons_ind_faixas_medicao[f - 1] + val, 2)

        # CALCULAR A TARIFA
        df_individual, valor_faixa1_conta = calcular_tarifa()
        #st.write(df_individual)

        # Valor da cota mínima individual
        # Dividir o valor do consumo individual alocado na faixa1 pelo número de apartamentos
        self.valor_cota_min_ind = round(valor_faixa1_conta / self.num_apartamentos, 2)
        self.exibir_resumo("Valor da cota mínima individual", self.valor_cota_min_ind)

        # Preparar o rateio
        df_rateio, df_resumo_final = preparar_rateio(df_individual)

        st.write(df_resumo_final)

        #st.write(df_rateio)

        # Exibir o valor final da conta
        self.exibir_resumo("Valor final da conta", round(df_rateio['valor_final'].sum(), 2))

        # Teste 16/06
        #st.write(df_individual)

        #st.write(exibir_detalhes_unidade(1304))

        st.success('Valor da conta --> R$ ' + real_br_money_mask(round(df_rateio['valor_final'].sum(), 2)))

        # Download do resultado em Excel com mais de uma aba - não utilizado no momento
        # list of dataframes
        # dfs = [df_resumo_final, df_rateio]

        # list of sheet names
        # sheets = ['resumo', 'rateio']
        # Gerar o Excel
        # df_xlsx = dfs_tabs(dfs, sheets, 'multi-test.xlsx')
        # Botão de Download
        # st.download_button(label='📥 Download', data=df_xlsx, file_name='df_test.xlsx')





        # Gerar o Excel com formatação
        df_form = gerar_excel_formatado(df_resumo_final, self.total_geral, self.tra)
        # Botão de Download
        st.download_button(label='Download',
                           data=df_form,
                           file_name='Quintessenza - Rateio Água.xlsx')

        self.unidade = st.number_input("Unidade", min_value=0, value=self.unidade, step=1)
        if self.unidade > 0:
            st.subheader(f"Detalhes da unidade {self.unidade}:")
            exibir_detalhes_unidade(self.unidade)







    def verif_arquivo(self):
        erro=[]
        # Verificar a quantidade de linhas do arquivo de entrada
        if len(self.entrada) != self.num_apartamentos:
            erro.append('Número de unidades incorreto')

        # Falta adicionar mais verificações

        # Número de erros
        if len(erro) == 0:
            # st.success('Planilha Excel verificada com sucesso.')

            # Chamar a próxima etapa para cálculo do rateio
            self.calcular_rateio()
        else:
            st.error(f"Arquivo com problemas. Número de erros: {len(erro)}")





def main():
    rateio = Rateio()

    st.title("Conta de água do condomínio Quintessenza")

    rateio.menu_lateral()

    rateio.pagina_principal()



if __name__ == "__main__":
    st.set_page_config(
        page_title="Rateio da conta de água", page_icon=":chart_with_upwards_trend:"
    )
    main()

